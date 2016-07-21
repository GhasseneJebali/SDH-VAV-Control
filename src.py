# -*- coding: utf-8 -*-
"""
Created on Wed May 04 16:06:20 2016

@author: Ghassene Jebali jbali.ghassen@gmail.com
"""

import openpyxl
import xlsxwriter
import os

###############################################################################
def write_output( vent, heat, setpt, date, debug):  
###############################################################################
    import openpyxl
    
    wb = openpyxl.load_workbook('Control.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    r = sheet.max_row

    
    sheet.cell(row=1, column=1).value = 'Ventillation'
    sheet.cell(row=r+1, column=1).value = vent
    
    sheet.cell(row=1, column=2).value = 'Heat'
    sheet.cell(row=r+1, column=2).value = heat
    
    sheet.cell(row=1, column=3).value = 'Set point'
    sheet.cell(row=r+1, column=3).value = setpt
    
    sheet.cell(row=1, column=4).value = 'Time'
    sheet.cell(row=r+1, column=4).value = date
    
    sheet.cell(row=1, column=5).value = 'Debug'
    sheet.cell(row=r+1, column=5).value = debug
    try:
        wb.save('Control.xlsx') 
    except Exception: 
        pass


###############################################################################
def setup():  
###############################################################################
    #import Data    
    import Prediction
    import warnings
    import Control
    
    warnings.filterwarnings("ignore")

    Prediction_horizon = 60 # in minutes

    #Data.data_acquisition()
    DATA_LIST={}
    wb = openpyxl.load_workbook('DATA_LIST.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    for key in range(1, sheet.max_column+1):
        DATA_LIST[sheet.cell(row=1, column=key).value]=[]
        for v in range(2, sheet.max_row+1):
            DATA_LIST[sheet.cell(row=1, column=key).value].append(sheet.cell(row=v, column=key).value)
     
    # Model generation
      
    #SVR_model = Prediction.Support_Vector_Regression(DATA_LIST, Prediction_horizon)
    KNN_model = Prediction.kNN_Regression(DATA_LIST, Prediction_horizon)
    BRR_model = Prediction.Bayesian_Ridge_Regression(DATA_LIST, Prediction_horizon)
    
    
#    workbook = xlsxwriter.Workbook(os.path.dirname(os.path.abspath(__file__))+'\Control.xlsx')
#    workbook.add_worksheet()
#    workbook.close()
    
    
    alpha = 0.7 # Higher alpha means slower adaptation
    T_history={}
    try:
        for i in range(1,8):
              T_history[str(i)] = Control.max_T_history(Control.previous_date(i)) 
        Mean_Running_Average =  (1.0 - alpha) * sum( (alpha** (i-1) ) * int(T_history [str(i)]) for i in range(1,8) )
    except Exception:
        Mean_Running_Average = 20

    
    return DATA_LIST, KNN_model, BRR_model, Mean_Running_Average
    
    
    
###############################################################################    
def update(d, model, client, state, area, Mean_Running_Average, debug):  
###############################################################################
    import Data
    import Prediction
    import Control
    import warnings
    import time, datetime  
    
    
    warnings.filterwarnings("ignore")
     
    date = datetime.datetime.fromtimestamp(time.time()).strftime('%Y-%m-%d %H:%M:%S')     
    warning = 0
    
    # Real time data
    try:    
        [T, co2, set_point, hum, T_outdoor, Human_date, Season, Cal_data, vent_power, H_C_power, cool_power, hour]=Data.Real_Time_Data(client)
    except Exception, e:
        print 'WARNING : PROBLEM DETECTED 1'
        print e
        warning = 3
        [T, co2, set_point, T_outdoor,   Cal_data,  H_C_power,  hour] = [23,400,23,23,2,-12,12]
        
    #occupancy prdiction
    try:
        state, Human_power, number = Prediction.occupancy(state, Cal_data, hour, co2)  
    except Exception, e:
        print 'WARNING : PROBLEM DETECTED 2'
        print e
        warning = 2
        state='occupied'
        Human_power=3.0
        number=30

    # Temperature prediction   
    T_needed_data=[T , H_C_power,  set_point, T_outdoor, Cal_data, Human_power]
    try: 
        T_predicted, warning= Prediction.T_prediciton(d, T_needed_data , model)
    except Exception, e:
        print 'WARNING : PROBLEM DETECTED 3'
        print e
        warning = 2
        T_predicted= T
    try:
          vent, heat, setpt, Mean_Running_Average = Control.control(state, number/6, area, T_outdoor, co2, T_predicted, Mean_Running_Average )
    except Exception, e:
        print Exception
        print 'WARNING : PROBLEM DETECTED 4'
        print e
        warning = 4
        heat = 0               
        setpt = 23
        vent = 200 
        
    setpt = round((setpt*9/5)+32,1)
    
    #write_output( vent, heat, setpt, date, debug)   

    
        
    
    print heat
    print setpt
    print vent
    print '---------------------------------------'
        
    return Mean_Running_Average, state, vent, heat, setpt, warning
