# -*- coding: utf-8 -*-
"""
Created on Wed May 04 16:06:20 2016

@author: Ghassene Jebali jbali.ghassen@gmail.com
"""

import time
import datetime
import warnings
import openpyxl
import prediction
import control
import data

def write_output(vent, heat, setpt, date, debug):
    wb = openpyxl.load_workbook("control.xlsx")
    sheet = wb.get_sheet_by_name("Sheet1")
    r = sheet.max_row

    sheet.cell(row=1, column=1).value = "Ventillation"
    sheet.cell(row=r + 1, column=1).value = vent

    sheet.cell(row=1, column=2).value = "Heat"
    sheet.cell(row=r + 1, column=2).value = heat

    sheet.cell(row=1, column=3).value = "Set point"
    sheet.cell(row=r + 1, column=3).value = setpt

    sheet.cell(row=1, column=4).value = "Time"
    sheet.cell(row=r + 1, column=4).value = date

    sheet.cell(row=1, column=5).value = "Debug"
    sheet.cell(row=r + 1, column=5).value = debug
    try:
        wb.save("control.xlsx")
    except Exception:
        pass


def setup():
    warnings.filterwarnings("ignore")

    prediction_horizon = 60  # in minutes

    # data.data_acquisition()
    DATA_LIST = {}
    wb = openpyxl.load_workbook("DATA_LIST.xlsx")
    sheet = wb.get_sheet_by_name("Sheet1")
    for key in range(1, sheet.max_column + 1):
        DATA_LIST[sheet.cell(row=1, column=key).value] = []
        for v in range(2, sheet.max_row + 1):
            DATA_LIST[sheet.cell(row=1, column=key).value].append(
                sheet.cell(row=v, column=key).value
            )

    # Model generation
    # SVR_model = prediction.Support_Vector_Regression(DATA_LIST, prediction_horizon)
    KNN_model = prediction.kNN_Regression(DATA_LIST, prediction_horizon)
    BRR_model = prediction.Bayesian_Ridge_Regression(DATA_LIST, prediction_horizon)

    #    workbook = xlsxwriter.Workbook(os.path.dirname(os.path.abspath(__file__))+'\control.xlsx')
    #    workbook.add_worksheet()
    #    workbook.close()

    alpha = 0.7  # Higher alpha means slower adaptation
    T_history = {}
    try:
        for i in range(1, 8):
            T_history[str(i)] = control.max_t_history(control.previous_date(i))
        mean_running_average = (1.0 - alpha) * sum(
            (alpha ** (i - 1)) * int(T_history[str(i)]) for i in range(1, 8)
        )
    except Exception:
        mean_running_average = 20

    return DATA_LIST, KNN_model, BRR_model, mean_running_average


def update(d, model, client, state, area, mean_running_average):
    warnings.filterwarnings("ignore")

    date = datetime.datetime.fromtimestamp(time.time()).strftime("%Y-%m-%d %H:%M:%S")
    warning = 0

    # Real time data
    try:
        [
            t,
            co2,
            set_point,
            hum,
            t_outdoor,
            human_date,
            season,
            cal_data,
            vent_power,
            h_c_power,
            cool_power,
            hour,
        ] = data.real_time_data(client)
    except Exception:
        warning = 3
        [t, co2, set_point, t_outdoor, cal_data, h_c_power, hour] = [
            23,
            401,
            23,
            23,
            2,
            -12,
            12,
        ]

    # occupancy prdiction
    try:
        state, human_power, number = prediction.occupancy(state, cal_data, hour, co2)
    except Exception:
        warning = 2
        state = "occupied"
        human_power = 3.0
        number = 30

    # Temperature prediction
    t_needed_data = [t, h_c_power, set_point, t_outdoor, cal_data, human_power]
    try:
        t_predicted, warning = prediction.t_prediciton(d, t_needed_data, model)
    except Exception:
        warning = 2
        t_predicted = t
    try:
        vent, heat, setpt, mean_running_average = control.control(
            state, number / 6, area, t_outdoor, co2, t_predicted, mean_running_average
        )
    except Exception:
        warning = 4
        heat = 0
        setpt = 23
        vent = 200

    setpt = round((setpt * 9 / 5) + 32, 1)

    return mean_running_average, state, vent, heat, setpt, warning
