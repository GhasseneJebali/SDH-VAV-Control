# -*- coding: utf-8 -*-
"""
Created on Wed May 18 16:16:20 2016

@author: ghassen
"""


from sklearn.svm import SVR
from sklearn import neighbors
from sklearn import linear_model
import warnings
import openpyxl
import xlsxwriter
import datetime



season = 'summer'

input_file = '\Simul_data_' + season + '.xlsx'
output_file = '\Simulation_' + season + '.xlsx'
weather_forecast_file = '\T_forecast_max_' + season + '.xlsx'
Running_average_file = '\Mean_Running_Average_' + season + '.xlsx'

###############################################################################
def Data_Validation(data, X):
###############################################################################
    valid=1    
    upper={}
    lower={}    
    for k in data.keys():
        upper[k]=max(data[k])
        lower[k]=min(data[k])
    if (X[0]<lower['Temperature']*0.9) or (X[0]>upper['Temperature']*1.1):
        valid=0
    if (X[1]<lower['H/C power']*0.9) or (X[1]>upper['H/C power']*1.1):
        valid=0
    if (X[2]<lower['Setpoint Temperature']*0.9) or (X[2]>upper['Setpoint Temperature']*1.1):
        valid=0
    if (X[3]<lower['Outdoor Temperature']*0.9) or (X[3]>upper['Outdoor Temperature']*1.1):
        valid=0
    if (X[4]<lower['Calendar data']*0.9) or (X[4]>upper['Calendar data']*1.1):
        valid=0
    return valid



###############################################################################
def Support_Vector_Regression(data, Prediction_horizon):  
###############################################################################
    # Import data

    X=[]
    y=[]
    X_train=[] 
    X_valid=[]
    y_train=[] 
    y_valid=[] 
    for i in range(len(data['Temperature'])-4):
        X.append([data['Temperature'][i] , data['H/C power'][i], data['Setpoint Temperature'][i], data['Outdoor Temperature'][i], data['Calendar data'][i], data['Human_power'][i]])
        y.append( data['Temperature'][i+4])
    X_train=X[:6000]
    X_valid=X[6000:]
    y_train=y[:6000]
    y_valid=y[6000:]

###############################################################################
    # Fit regression model
    svr_rbf = SVR(kernel='rbf', C=1e3, gamma=0.1)

    y_rbf = svr_rbf.fit(X_train, y_train)
    
    result=[]
    error=[]
    for i in range(len(X_valid)):
        result.append(y_rbf.predict(X_valid[i]))
        error.append(abs(result[i]-y_valid[i]))
        
    return y_rbf
###############################################################################




###############################################################################
def kNN_Regression(data, Prediction_horizon): 
###############################################################################
    # Import data

    X=[]
    y=[]
    X_train=[] 
    X_valid=[]
    y_train=[] 
    y_valid=[] 
    for i in range(len(data['Temperature'])-4):
        X.append([data['Temperature'][i] , data['H/C power'][i], data['Setpoint Temperature'][i], data['Outdoor Temperature'][i], data['Calendar data'][i], data['Human_power'][i]])
        y.append( data['Temperature'][i+4])
    X_train=X[:6000]
    X_valid=X[6000:]
    y_train=y[:6000]
    y_valid=y[6000:]
###############################################################################
    # Fit regression model
    n_neighbors = 5
    knn = neighbors.KNeighborsRegressor(n_neighbors, weights= 'distance')
    
    y_kNN = knn.fit(X_train, y_train)
    
    result=[]
    error=[]
    for i in range(len(X_valid)):
        result.append(y_kNN.predict(X_valid[i]))
        error.append(abs(result[i]-y_valid[i]))
        
    return y_kNN
###############################################################################


###############################################################################
def Bayesian_Ridge_Regression(data, Prediction_horizon): 
###############################################################################
    # Import data

    X=[]
    y=[]
    X_train=[] 
    X_valid=[]
    y_train=[] 
    y_valid=[] 
    for i in range(len(data['Temperature'])-4):
        X.append([data['Temperature'][i] , data['H/C power'][i], data['Setpoint Temperature'][i], data['Outdoor Temperature'][i], data['Calendar data'][i], data['Human_power'][i]])
        y.append( data['Temperature'][i+4])
    X_train=X[:6000]
    X_valid=X[6000:]
    y_train=y[:6000]
    y_valid=y[6000:]
###############################################################################
    # Fit regression model
    clf = linear_model.BayesianRidge()
    
    clf.fit(X_train, y_train)
    
    result=[]
    error=[]
    for i in range(len(X_valid)):
        result.append(clf.predict (X_valid[i]))
        error.append(abs(result[i]-y_valid[i]))
        
    return clf
###############################################################################


###############################################################################
def T_prediciton(data, X , model):    
###############################################################################  
    if Data_Validation(data, X):
        T = model.predict(X)[0]
    else:
        print 'input data is invalid or may be outside boundaries'
        T = X[0]
    return T
###############################################################################


###############################################################################
def occupancy(state, cal, h, co2):
###############################################################################
    if state=='start':
        if cal==2 and h<17 and h>=10:
            state='occupied'
            Human_power=3.0
            number=30
            return state, Human_power, number
        if cal==2 and ((h<10 and h>=7) or (h<19 and h>=17)):
            state='slightly occupied'
            Human_power=1.5
            number=15
            return state, Human_power, number
        else:
            state='not occupied'
            Human_power=0.0
            number=0
            return state, Human_power, number
    if state=='not occupied':
        if cal==1 or cal==0:
            state='not occupied'
            Human_power=0.0
            number=0
            return state, Human_power, number
        else:
            state='slightly occupied'
            Human_power=1.5
            number=15
            return state, Human_power, number
    if state=='slightly occupied':
        if (cal==2 and h<17 and h>=10) or co2>500:
            state='occupied'
            Human_power=3.0
            number=30
            return state, Human_power, number
        if cal==0 or cal==1:
            state='not occupied'
            Human_power=0.0
            number=0
            return state, Human_power, number
        else:
            state='slightly occupied'
            Human_power=1.5
            number=15
            return state, Human_power, number
    if state=='occupied':
        if (cal==2 and h<17 and h>=10) or co2>500:
            state='occupied'
            Human_power=3.0
            number=30
            return state, Human_power, number
        else:
            state='slightly occupied'
            Human_power=1.5
            number=15
            return state, Human_power, number
            
###############################################################################
# returns the mean temperature of a given date
def max_T_history(date):
    import requests
    
    urlstart = 'http://api.wunderground.com/api/a34c4ac816767601/history_'
    urlend = '/q/CA/berkeley.json'
    url = urlstart + date + urlend
    data = requests.get(url).json()
    #data['history']['dailysummary'][0]['meantempm']
    return data['history']['dailysummary'][0]['maxtempm']
    

# returns the date of a previous delta day
def previous_date(Timestamp, delta):
    
    import time
    import datetime
    
    date = datetime.datetime.fromtimestamp(Timestamp).strftime('%Y-%m-%d %H:%M:%S')
    year = int (date[0:4])
    month = int (date[5:7])
    day = int (date[8:10])

    day_of_the_year = time.strptime(str(day)+' '+str(month)+' '+str(year), "%d %m %Y")[7]
    
    if int(day_of_the_year) > int(delta):
        day = int(day_of_the_year)-int(delta)
    else:
        day = 366 + int(time.strftime("%j"))-int(delta)
        year = year-1
        
    now = str(year)+' '+str(day)
    date = str(datetime.datetime.strptime(now, '%Y %j'))
    y = str(date[0:4])
    m = str(date[5:7])
    d = str(date[8:10])
    
   
    
    return y+m+d
 
# returns the minimum and maximum temperature for a given running average. 90% acceptability
def ACZ(Timestamp, Mean_Running_Average):    
#    alpha = 0.7 # Higher alpha means slower adaptation
#    T_history={}
#    
#    date = datetime.datetime.fromtimestamp(Timestamp).strftime('%Y-%m-%d %H:%M:%S')
#    hour = int(date[11:13])
    
#    if hour == 0:
#        for i in range(1,8):
#            T_history[str(i)] = int(max_T_history(previous_date(Timestamp, i)))
#        Mean_Running_Average =  (1.0 - alpha) * sum( (alpha** (i-1) ) * int(T_history [str(i)]) for i in range(1,8) )
    
    Upper_limit = 0.31 * Mean_Running_Average + 20.3
    Lower_limit = 0.31 * Mean_Running_Average + 15.3
    
    return Lower_limit, Upper_limit
    

def control(state, N_person, area, T_outdoor, co2, T_predicted, hour, Timestamp,  Mean_Running_Average, T_forecast_max ):
#    import requests
      
    
    date = datetime.datetime.fromtimestamp(Timestamp).strftime('%Y-%m-%d %H:%M:%S')
#    year = str (date[0:4])
    month = str (date[5:7])
    day = str (date[8:10])
    hour = int(date[11:13])
#    urlstart = 'http://api.wunderground.com/api/f6c680e843927e80/history_'
#    urlend = '/q/CA/San_Francisco.json'
#    url = urlstart + year + month + day + urlend
#    Hum_actual = requests.get(url).json()['history']['observations'][hour]['hum']
    
    
    
    #T_forecast_max = int(max_T_history(previous_date(Timestamp, 0)) )
    
#    w1 = openpyxl.load_workbook('C:\Users\ghassen\Dropbox\Mastere OSE\Stage\Python Scripts\Results\T_Forecast_max_winter.xlsx')
#    sheet1 = w1.get_sheet_by_name('Sheet1')
#    r1 = sheet1.max_row
#    sheet1.cell(row=1, column=5).value = 'T_max_forecast'
#    sheet1.cell(row=r1+1, column=5).value = T_forecast_max    
#    sheet1.cell(row=1, column=1).value = 'Time'
#    sheet1.cell(row=r1+1, column=1).value = Timestamp   
#    sheet1.cell(row=1, column=2).value = 'Day'
#    sheet1.cell(row=r1+1, column=2).value = day
#    sheet1.cell(row=1, column=3).value = 'Month'
#    sheet1.cell(row=r1+1, column=3).value = month
#    sheet1.cell(row=1, column=4).value = 'hour'
#    sheet1.cell(row=r1+1, column=4).value = hour
#    w1.save('C:\Users\ghassen\Dropbox\Mastere OSE\Stage\Python Scripts\Results\T_Forecast_max_winter.xlsx') 
#    
    # temperature setpoints
    
    Upper_T_limit = 0.31 * Mean_Running_Average + 20.3
    Lower_T_limit = 0.31 * Mean_Running_Average + 15.3
    
#    w2 = openpyxl.load_workbook('C:\Users\ghassen\Dropbox\Mastere OSE\Stage\Python Scripts\Results\Mean_Running_Average_winter.xlsx')
#    sheet2 = w2.get_sheet_by_name('Sheet1')
#    r2 = sheet2.max_row
#    sheet2.cell(row=1, column=5).value = 'Mean'
#    sheet2.cell(row=r2+1, column=5).value = Mean_Running_Average    
#    sheet2.cell(row=1, column=1).value = 'Time'
#    sheet2.cell(row=r2+1, column=1).value = Timestamp   
#    sheet2.cell(row=1, column=2).value = 'Day'
#    sheet2.cell(row=r2+1, column=2).value = day
#    sheet2.cell(row=1, column=4).value = 'hour'
#    sheet2.cell(row=r2+1, column=4).value = hour
#    sheet2.cell(row=1, column=3).value = 'Month'
#    sheet2.cell(row=r2+1, column=3).value = month
#    w2.save('C:\Users\ghassen\Dropbox\Mastere OSE\Stage\Python Scripts\Results\Mean_Running_Average_winter.xlsx')     
    
    
    Hight_setpt = (Lower_T_limit + 3*Upper_T_limit)/4
    Center_setpt = (Lower_T_limit + Upper_T_limit)/2
    Low_setpt = (3*Lower_T_limit + Upper_T_limit)/4
    Hight_setpt=round(Hight_setpt,1)
    Center_setpt=round(Center_setpt,1)
    Low_setpt=round(Low_setpt,1)
    
    
    # ventillation setpoint
    CA = 0.06 # cfm/ft2
    CP = 5 # cfm/person
    Zone_Air_Distribution_Effectiveness = 1.0
    Vent_min = (CA*area + CP*N_person) / Zone_Air_Distribution_Effectiveness
    Vent_setpt = 2*Vent_min


    # morning afternnon determination
    if hour < 12:
        AM_PM = 'AM'
    else:
        AM_PM = 'PM'     
      
        
    if state == 'not occupied':
        if Mean_Running_Average > 15 :
           
            vent = Vent_setpt*2
            heat = 0
            setpt = Lower_T_limit
            return   vent, heat, setpt, Mean_Running_Average
        else:
           
            vent = Vent_setpt*2
            heat = 1
            setpt = Lower_T_limit
            return    vent, heat, setpt, Mean_Running_Average
                   
    if state == 'occupied':
        if co2 > 800:
      
            vent = Vent_setpt*4
            heat = 0
            setpt = Center_setpt
            return   vent, heat, setpt, Mean_Running_Average
        if co2 <= 800:
            if T_predicted > Upper_T_limit:
              
                vent = Vent_setpt*4
                heat = 0
                setpt = Center_setpt
                return  vent, heat, setpt, Mean_Running_Average
            if T_predicted < Lower_T_limit:
              
                vent = Vent_setpt*3
                heat = 1
                setpt = Low_setpt
                return  vent, heat, setpt, Mean_Running_Average
            else:
                if Mean_Running_Average <= 15:
                    if AM_PM == 'AM':
                      
                        vent = Vent_setpt*3
                        heat = 1
                        setpt = Low_setpt
                        return   vent, heat, setpt, Mean_Running_Average
                    if AM_PM == 'PM':
                      
                        vent = Vent_setpt*3
                        heat = 1
                        setpt = Center_setpt
                        return   vent, heat, setpt, Mean_Running_Average
                if Mean_Running_Average > 15:
                    if AM_PM == 'AM':
                       
                        vent = Vent_setpt*3
                        heat = 0
                        setpt = Low_setpt
                        return   vent, heat, setpt, Mean_Running_Average
                    if AM_PM == 'PM':
                      
                        vent = Vent_setpt*3
                        heat = 0
                        setpt = Center_setpt
                        return   vent, heat, setpt , Mean_Running_Average    

    if state == 'slightly occupied':
        if ( (AM_PM == 'PM') or ( (AM_PM == 'AM') and (T_forecast_max < 28) ) ):
            
            if Mean_Running_Average > 15:
                if T_predicted > Upper_T_limit:
          
                    vent = Vent_setpt*3
                    heat = 0
                    setpt = Center_setpt
                    return  vent, heat, setpt, Mean_Running_Average
                if T_predicted < Lower_T_limit:
                
                    vent = Vent_setpt*2
                    heat = 0
                    setpt = Low_setpt
                    return  vent, heat, setpt, Mean_Running_Average
                else:
                 
                    vent = Vent_setpt*2
                    heat = 0
                    setpt = Center_setpt
                    return  vent, heat, setpt, Mean_Running_Average
            if Mean_Running_Average <= 15:
                if T_predicted > Upper_T_limit:
               
                    vent = Vent_setpt*3
                    heat = 0
                    setpt = Center_setpt
                    return  vent, heat, setpt, Mean_Running_Average
                if T_predicted < Lower_T_limit:

                    vent = Vent_setpt*2
                    heat = 1
                    setpt = Low_setpt
                    return   vent, heat, setpt, Mean_Running_Average
                else:

                    vent = Vent_setpt*2
                    heat = 1
                    setpt = Center_setpt
                    return  vent, heat, setpt, Mean_Running_Average
        else:
            vent = Vent_setpt*3
            heat = 0
            setpt = Center_setpt
            return   vent, heat, setpt, Mean_Running_Average
    
warnings.filterwarnings("ignore")

Prediction_horizon = 60 # in minutes

#Data.data_acquisition()
DATA_LIST={}
wb = openpyxl.load_workbook('C:\Users\ghassen\Dropbox\Mastere OSE\Stage\Python Scripts\Results\Fixed'+ input_file)
sheet = wb.get_sheet_by_name('Sheet1')
for key in range(1, sheet.max_column+1):
    DATA_LIST[sheet.cell(row=1, column=key).value]=[]
    for v in range(2, sheet.max_row+1):
        DATA_LIST[sheet.cell(row=1, column=key).value].append(sheet.cell(row=v, column=key).value)

Mean_Running_Average_dic={}
w1 = openpyxl.load_workbook('C:\Users\ghassen\Dropbox\Mastere OSE\Stage\Python Scripts\Results\Fixed' + Running_average_file)
sheet1 = w1.get_sheet_by_name('Sheet1')
for key in range(1, sheet1.max_column+1):
    Mean_Running_Average_dic[sheet1.cell(row=1, column=key).value]=[]
    for v in range(2, sheet1.max_row+1):
        Mean_Running_Average_dic[sheet1.cell(row=1, column=key).value].append(sheet1.cell(row=v, column=key).value)

T_forecast_max_dic={}
w2 = openpyxl.load_workbook('C:\Users\ghassen\Dropbox\Mastere OSE\Stage\Python Scripts\Results\Fixed' + weather_forecast_file)
sheet2 = w2.get_sheet_by_name('Sheet1')
for key in range(1, sheet2.max_column+1):
    T_forecast_max_dic[sheet2.cell(row=1, column=key).value]=[]
    for v in range(2, sheet2.max_row+1):
        T_forecast_max_dic[sheet2.cell(row=1, column=key).value].append(sheet2.cell(row=v, column=key).value)


  
#SVR_model = Prediction.Support_Vector_Regression(DATA_LIST, Prediction_horizon)
#KNN_model = kNN_Regression(DATA_LIST, Prediction_horizon)
BRR_model = Bayesian_Ridge_Regression(DATA_LIST, Prediction_horizon)

workbook = xlsxwriter.Workbook('C:\Users\ghassen\Dropbox\Mastere OSE\Stage\Python Scripts\Results'+ output_file)
workbook.add_worksheet()
workbook.close()

#workbook2 = xlsxwriter.Workbook('C:\Users\ghassen\Dropbox\Mastere OSE\Stage\Python Scripts\Results\Mean_Running_Average_winter.xlsx')
#workbook2.add_worksheet()
#workbook2.close()
#
#workbook3 = xlsxwriter.Workbook('C:\Users\ghassen\Dropbox\Mastere OSE\Stage\Python Scripts\Results\T_Forecast_max_winter.xlsx')
#workbook3.add_worksheet()
#workbook3.close()

state = 'start'

Ventillataion=[]
Heating=[]
Setpoint=[]
Time=[]
#Mean_Running_Average = 20
for i in range(len(Mean_Running_Average_dic['Mean'])):

    # Real time data
    T = DATA_LIST['Temperature'][i]   
    co2 = DATA_LIST['CO2'][i] 
    T_outdoor = DATA_LIST['Outdoor Temperature'][i]   
    set_point = DATA_LIST['Setpoint Temperature'][i]   
    Cal_data = DATA_LIST['Calendar data'][i]   
    H_C_power = DATA_LIST['H/C power'][i]   
    Timestamp = DATA_LIST['Timestamp'][i]/1000 
    
    Mean_Running_Average = Mean_Running_Average_dic['Mean'][i]   
    T_forecast_max = T_forecast_max_dic['T_max_forecast'][i]  
    
    date = datetime.datetime.fromtimestamp(Timestamp).strftime('%Y-%m-%d %H:%M:%S')
    hour = int(date[11:13]) 

    #occupancy prdiction
    state, Human_power, number = occupancy(state, Cal_data, hour, co2)  
    
    # Temperature prediction
    T_needed_data=[T , H_C_power,  set_point, T_outdoor, Cal_data, Human_power]
    T_predicted= T_prediciton(DATA_LIST, T_needed_data , BRR_model)


    vent, heat, setpt, Mean_Running_Average= control(state, number/6, 296, T_outdoor, co2, T_predicted, hour, Timestamp, Mean_Running_Average, T_forecast_max )
    
    Ventillataion.append(vent)
    Heating.append(heat)
    Setpoint.append(setpt)
    Time.append(date)
    
wb = openpyxl.load_workbook('C:\Users\ghassen\Dropbox\Mastere OSE\Stage\Python Scripts\Results'+ output_file)
sheet = wb.get_sheet_by_name('Sheet1')
r = sheet.max_row

sheet.cell(row=1, column=1).value = 'Ventillation'
sheet.cell(row=1, column=2).value = 'Heat'
sheet.cell(row=1, column=3).value = 'Set point'
sheet.cell(row=1, column=4).value = 'Time'

for i in range(len(Ventillataion)):
    #print round(100 * float(i)/float(len(Ventillataion)) ,1),'%'
     
    sheet.cell(row=r+1, column=1).value = Ventillataion[i]

    sheet.cell(row=r+1, column=2).value = Heating[i]
    
    sheet.cell(row=r+1, column=3).value = Setpoint[i]
    
    sheet.cell(row=r+1, column=4).value = Time[i]
    
    r=r+1
    
wb.save('C:\Users\ghassen\Dropbox\Mastere OSE\Stage\Python Scripts\Results' + output_file)  

print 'Done'
